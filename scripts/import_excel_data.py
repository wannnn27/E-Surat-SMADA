"""Validasi dan impor data master Excel ke JSON aplikasi.

Secara default perintah hanya memeriksa data. Tambahkan ``--write`` setelah
laporan validasi bersih untuk mengganti ketiga JSON secara atomik.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_ROOT = Path(os.getenv("ESURAT_DATA_ROOT", str(BASE_DIR / "data"))).expanduser()
MASTER_DATA_DIR = Path(os.getenv("ESURAT_DATA_DIR", str(DATA_ROOT / "master"))).expanduser()
EXCEL_DIR = Path(os.getenv("ESURAT_SOURCE_DIR", str(DATA_ROOT / "source"))).expanduser()

DEFAULT_FILES = {
    "guru": "nomonatif guru peg sma2 2026.xlsx",
    "murid": "2627=daftar_murid=v1.3.xlsx",
    "kode": "KODE KLASIFIKASI ARSIP-SMAN 2 WONOSARI 2025.xlsx",
}
MAX_EXACT_INTEGER_FLOAT = (1 << 53) - 1


class ImportValidationError(ValueError):
    """Kesalahan data master yang harus diperbaiki sebelum publikasi."""


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_identifier(value: Any, *, label: str, length: int | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        raise ImportValidationError(f"{label} tidak boleh berupa nilai boolean")
    if isinstance(value, int):
        result = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            raise ImportValidationError(f"{label} harus berupa bilangan bulat: {value!r}")
        if abs(value) > MAX_EXACT_INTEGER_FLOAT:
            raise ImportValidationError(
                f"{label} terlalu panjang untuk sel angka Excel; format kolom sebagai teks"
            )
        result = str(int(value))
    else:
        result = clean_text(value)
        if re.fullmatch(r"\d+\.0", result):
            result = result[:-2]

    if result and not result.isdigit():
        raise ImportValidationError(f"{label} harus berisi angka saja: {result!r}")
    if result and length is not None and len(result) != length:
        raise ImportValidationError(
            f"{label} harus {length} digit, ditemukan {len(result)} digit: {result!r}"
        )
    return result


def ensure_unique(rows: Iterable[dict[str, str]], field: str, label: str) -> None:
    values = [row[field] for row in rows]
    duplicates = [value for value, count in Counter(values).items() if count > 1]
    if duplicates:
        sample = ", ".join(duplicates[:5])
        raise ImportValidationError(f"{label} duplikat ditemukan: {sample}")


def require_fields(rows: Iterable[dict[str, str]], fields: Iterable[str], label: str) -> None:
    for index, row in enumerate(rows, start=1):
        missing = [field for field in fields if not clean_text(row.get(field))]
        if missing:
            raise ImportValidationError(
                f"{label} baris data #{index} tidak memiliki: {', '.join(missing)}"
            )


def split_kelas(raw_value: Any) -> tuple[str, str]:
    raw = clean_text(raw_value)
    parts = re.split(r",\s*mapel\s+pilihan\s*:\s*", raw, maxsplit=1, flags=re.IGNORECASE)
    kelas = parts[0].strip().upper()
    mapel_pilihan = parts[1].strip() if len(parts) == 2 else ""
    if not re.fullmatch(r"(?:X|XI|XII)-[A-Z]", kelas):
        raise ImportValidationError(f"Format kelas tidak dikenali: {raw!r}")
    return kelas, mapel_pilihan


def import_guru(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows: list[dict[str, str]] = []
        for excel_row, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if excel_row == 1 or len(row) < 10:
                continue
            try:
                nip = normalize_identifier(row[1], label=f"NIP baris Excel {excel_row}", length=18)
            except ImportValidationError:
                if not clean_text(row[1]):
                    continue
                raise
            if not nip:
                continue
            rows.append(
                {
                    "nip": nip,
                    "nama": clean_text(row[2]),
                    "ttl": clean_text(row[3]),
                    "golongan": clean_text(row[5]),
                    "tmt_golongan": clean_text(row[6]),
                    "jabatan": clean_text(row[7]),
                    "status_pegawai": clean_text(row[8]),
                    "kedudukan": clean_text(row[9]) or "Aktif",
                }
            )
    finally:
        workbook.close()

    require_fields(rows, ("nip", "nama", "jabatan", "golongan"), "Guru")
    ensure_unique(rows, "nip", "NIP")
    if not rows:
        raise ImportValidationError("Tidak ada data guru valid yang ditemukan")
    return rows


def import_murid(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows: list[dict[str, str]] = []
        target_sheets = [name for name in ("10", "11", "12") if name in workbook.sheetnames]
        if not target_sheets:
            raise ImportValidationError("Sheet siswa 10, 11, atau 12 tidak ditemukan")

        for sheet_name in target_sheets:
            sheet = workbook[sheet_name]
            current_kelas = ""
            current_mapel = ""
            for excel_row, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                first_cell = clean_text(row[0]) if row else ""
                if re.search(r"kelas\s*:", first_cell, flags=re.IGNORECASE):
                    raw_kelas = re.split(r"kelas\s*:\s*", first_cell, maxsplit=1, flags=re.IGNORECASE)[-1]
                    current_kelas, current_mapel = split_kelas(raw_kelas)
                    continue

                if len(row) < 4 or not clean_text(row[3]):
                    continue
                try:
                    sequence = normalize_identifier(row[0], label=f"Nomor urut {sheet_name}!{excel_row}")
                except ImportValidationError:
                    continue
                if not sequence:
                    continue
                if not current_kelas:
                    raise ImportValidationError(
                        f"Data siswa {sheet_name}!{excel_row} muncul sebelum header kelas"
                    )

                nis = normalize_identifier(row[1], label=f"NIS {sheet_name}!{excel_row}")
                nisn = normalize_identifier(row[2], label=f"NISN {sheet_name}!{excel_row}", length=10)
                rows.append(
                    {
                        "nis": nis,
                        "nisn": nisn,
                        "nama": clean_text(row[3]),
                        "jk": clean_text(row[4]) if len(row) > 4 else "",
                        "agama": clean_text(row[5]) if len(row) > 5 else "",
                        "kelas": current_kelas,
                        "mapel_pilihan": current_mapel,
                    }
                )
    finally:
        workbook.close()

    require_fields(rows, ("nis", "nisn", "nama", "kelas"), "Siswa")
    ensure_unique(rows, "nis", "NIS")
    ensure_unique(rows, "nisn", "NISN")
    if not rows:
        raise ImportValidationError("Tidak ada data siswa valid yang ditemukan")
    return rows


def import_kode_arsip(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows: list[dict[str, str]] = []
        for excel_row, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if len(row) < 3 or not clean_text(row[1]) or not clean_text(row[2]):
                continue
            kode = clean_text(row[1])
            if not re.fullmatch(r"\d+(?:\.\d+)*", kode):
                if excel_row <= 3:
                    continue
                raise ImportValidationError(f"Kode arsip tidak valid pada baris {excel_row}: {kode!r}")
            rows.append({"kode": kode, "keterangan": clean_text(row[2])})
    finally:
        workbook.close()

    require_fields(rows, ("kode", "keterangan"), "Kode arsip")
    ensure_unique(rows, "kode", "Kode arsip")
    if not rows:
        raise ImportValidationError("Tidak ada kode arsip valid yang ditemukan")
    return rows


def resolve_input(explicit: str | None, default_name: str, keyword: str) -> Path:
    if explicit:
        path = Path(explicit).expanduser().resolve()
        if not path.exists():
            raise ImportValidationError(f"File tidak ditemukan: {path}")
        return path

    default_path = EXCEL_DIR / default_name
    if default_path.exists():
        return default_path

    matches = [
        path for path in EXCEL_DIR.glob("*.xlsx")
        if not path.name.startswith("~$") and keyword.casefold() in path.name.casefold()
    ]
    if len(matches) != 1:
        names = ", ".join(path.name for path in matches) or "tidak ada"
        raise ImportValidationError(
            f"Tidak dapat menentukan file {keyword!r} secara otomatis; kandidat: {names}"
        )
    return matches[0]


def write_json_atomic(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temp_name = tempfile.mkstemp(prefix=f".{path.stem}-", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(handle, "w", encoding="utf-8", newline="\n") as stream:
            json.dump(rows, stream, indent=2, ensure_ascii=False)
            stream.write("\n")
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temp_name, path)
    except Exception:
        try:
            os.unlink(temp_name)
        except FileNotFoundError:
            pass
        raise


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--guru-file", help="Path file Excel guru")
    parser.add_argument("--murid-file", help="Path file Excel siswa")
    parser.add_argument("--kode-file", help="Path file Excel kode klasifikasi")
    parser.add_argument(
        "--write",
        action="store_true",
        help="Tulis JSON setelah semua dataset lolos validasi (default: check-only)",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    guru_path = resolve_input(args.guru_file, DEFAULT_FILES["guru"], "guru")
    murid_path = resolve_input(args.murid_file, DEFAULT_FILES["murid"], "murid")
    kode_path = resolve_input(args.kode_file, DEFAULT_FILES["kode"], "klasifikasi")

    guru = import_guru(guru_path)
    murid = import_murid(murid_path)
    kode = import_kode_arsip(kode_path)

    print(f"[OK] Guru/staff : {len(guru)} record; NIP unik dan valid.")
    print(f"[OK] Siswa      : {len(murid)} record; NIS/NISN unik dan valid.")
    print(f"[OK] Kode arsip : {len(kode)} record unik.")
    print(f"[OK] Kelas      : {len({row['kelas'] for row in murid})} rombel; mapel pilihan dipisahkan.")

    if not args.write:
        print("[CHECK-ONLY] Tidak ada file yang diubah. Jalankan ulang dengan --write untuk publikasi.")
        return 0

    # Seluruh data sudah tervalidasi sebelum satu pun file tujuan diganti.
    write_json_atomic(MASTER_DATA_DIR / "guru.json", guru)
    write_json_atomic(MASTER_DATA_DIR / "murid.json", murid)
    write_json_atomic(MASTER_DATA_DIR / "kode_arsip.json", kode)
    print("[WRITE] Ketiga JSON berhasil diperbarui secara atomik per file.")
    print("[RESTART] Restart aplikasi agar proses aktif memuat data versi baru.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ImportValidationError as exc:
        print(f"[GAGAL] {exc}")
        raise SystemExit(2)
